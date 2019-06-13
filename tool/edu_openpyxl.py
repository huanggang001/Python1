from openpyxl import load_workbook
import os
import sys


# -------------------------------------------------------------------------------
# 类名称：getTestData
# 类的目的：获取测试数据
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：老道
# 创建时间：2019/05/18
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------
case_local = os.path.abspath(os.path.join(os.path.dirname(__file__), '../cases/EDU_V1.0测试用例.xlsx'))
login_requirename = 'Login'
case_stuff = 2

# 获取测试数据

def getTestData(local, requirement, casename):
    wb = load_workbook(local)
    ws = wb[requirement]
    for i in range(len(ws['A'])):
        if ws['A'][i].value == casename:
            result = ws.cell(row=ws['A'][i].row,column=5).value
            if result is not None:
                stuff = result.find(',')
                if stuff != -1:
                    data = result.split(',')
                    print('第一',len(data))
                    if " '" in data:
                        data =[None,data[1]]
                        print(len(data))
                        print('2',data[1])

                        return data
                    return data

    wb.close()
